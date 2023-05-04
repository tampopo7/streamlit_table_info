# Import required Libraries
# streamlit
import streamlit as st
# Snowpark
from snowflake.snowpark.session import Session
# Pandas
import pandas as pd
# excel
from openpyxl import reader, load_workbook, Workbook
# etc
import io
#from io import BytesIO
import json
#import datetime


def main():
    st.title('Snowflake Table List')
    st.write('Displays a list of tables for a given schema and downloads detailed information to an Excel file.')

    #ログイン情報取得
    con_param = get_login_info()
    if con_param:
        if st.button('Show Tables', on_click=None):

            # テーブル一覧取得
            session = Session.builder.configs(con_param).create()
            tb_list = session.sql("SHOW TABLES").collect()

            # テーブル一覧表示
            st.write(tb_list)

            # ブックを作成
            wb = Workbook()

            # テーブル一覧作成
            sheet = wb['Sheet']
            sheet.title = 'list'

            # テーブル名リスト
            tb_name_list = []

            head = True
            for row in tb_list:

                # RowType → Dict
                a = row.as_dict()

                # Head出力
                if head:
                    c = list(a.keys())
                    sheet.append(c)
                    head = False

                #Excel does not support timezones in datetimes. 
                a['created_on'] = a['created_on'].isoformat()
                # 辞書の値のみにする Noneを空白に
                #c = ['' if n is None else n for n in a.values()]
                v = ['' if n is None else n for n in list(a.values())]
                # 1行追加
                sheet.append(v)

                # テーブル名リストに追加
                tb_name_list.append(a['name'])

            database = con_param['database']
            schema = con_param['schema']
            for tb in tb_name_list:

                #１テーブルずつの処理
                cl_list = session.sql(f"SHOW COLUMNS IN TABLE {database}.{schema}.{tb}").collect()
                sh = wb.create_sheet(title=tb)

                head = True
                for row in cl_list:
                    # 1カラムずつ １行ずつの処理
                    # RowType → Dict
                    a = row.as_dict()

                    # Head出力
                    if head:
                        h = []
                        for k, v in a.items():
                            # data_typeは値が辞書型の文字列
                            if k == 'data_type':
                                # カラムの型によってdata_type列の出力項目が違うため、固定で持つ
                                #dic = json.loads(v)
                                #h.extend(list(dic.keys()))
                                h.extend(['length','byteLength','nullable','fixed'])
                            else:
                                h.append(k)
                        sh.append(h)
                        head = False

                    d = []
                    for k, v in a.items():
                        # 1カラムの中の１項目ずつの処理
                        # data_typeは値が辞書型の文字列
                        if k == 'data_type':
                            dic = json.loads(v)
                            #d.extend(list(dic.values()))
                            d.append(dic.get('length',''))
                            d.append(dic.get('byteLength',''))
                            d.append(dic.get('nullable',''))
                            d.append(dic.get('fixed',''))
                        else:
                            d.append(v)

                    # 辞書の値のみにする Noneを空白に
                    #c = ['' if n is None else n for n in d.values()]
                    # 1行追加
                    sh.append(d)

            session.close()

            # ブックをバッファに保存
            buffer=io.BytesIO()
            wb.save(buffer)

            st.download_button(
                label='Download table details to Excel'
                ,data=buffer
                ,file_name='snowflake_table.xlsx'
            )

#ログイン情報取得
def get_login_info():
    st.write('Please enter your login information')

    account = st.text_input('account', '')
    user = st.text_input('user', '')
    password = st.text_input('password', '')
    role = st.text_input('role', '')
    warehouse = st.text_input('warehouse', '')
    database = st.text_input('database', '')
    schema = st.text_input('schema', '')

    if account and user and password and role and warehouse and database and schema:
        connection_parameters = {
        "account": account,
        "user": user,
        "password": password,
        "role": role,
        "warehouse": warehouse,
        "database": database,
        "schema": schema 
        }
        return connection_parameters

main()
